from pathlib import Path
import json
import jsonpath
import openpyxl


val = Path(__file__).parent.parent
print(val)
print(type(val))
val1 = str(val)
print(val1)
print(type(val1))

# A function to read JSON value from the JSON file using json package and jsonpath package
def read_locator_from_json(locatorName):
    # locatorName is the KEY in ElementsRediffmail.json whose VALUE will be picked by this function
    # First open the JSON file
    f = open(str(val)+'\\JsonFiles\\ElementsRediffmail.json')
    # The file is read with read() method and then it will be parsed with loads() method of JSON package
    response = json.loads(f.read())

    # Travel to get the VAYE of the KEY defined in ElementsRediffmail.json JSON file
    value  = jsonpath.jsonpath(response, locatorName)
    # Jsonpath() method returns a list
    return value[0]

# Define the path of the excel sheet and load the workbook
workBooKPath = str(val)+"\\Resources\\TestData.xlsx"
wb = openpyxl.load_workbook(workBooKPath)



# A function to find maximum rows in an excel sheet
def get_max_row_number(sheetName):
    sh = wb[sheetName]
    maxRow = sh.max_row
    return maxRow

# A function to find maximum rows in an excel sheet
def get_max_column_number(sheetName):
    sh = wb[sheetName]
    maxColumn = sh.max_column
    return maxColumn

# Function to get the data from a cell  by its row number and column number
def get_data_from_cell(sheetName, row, column):
    sh = wb[sheetName]
    # To defy the problem that row and column is taken as int value, we will deliberately typecase to int
    data = sh.cell(int(row), int(column));
    return data.value



maxR = get_max_row_number("Data")
print("Maximum number of rows filled with data are",maxR)

maxC = get_max_column_number("Data")
print("Maximum number of columns filled with data are",maxC)


dataCell1= get_data_from_cell("Data", 2, 2)
print("Data in second row and second column is",dataCell1)

dataCell2= get_data_from_cell("Data", 3, 5)
print("Data in third row and fifth column is",dataCell2)


